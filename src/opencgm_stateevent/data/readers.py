"""Minimal raw readers, one per source.

These do the smallest amount of work needed to yield ``(local_datetime, glucose_mg_dl)`` with
its source identity. No resampling, no interpolation, no unit guessing beyond what each source
documents. The canonical Parquet layer (blueprint §8.2) is built on top of these.

Every reader preserves the source's own local wall-clock time. Absolute time-of-day is a model
input (§8.5), so converting to UTC here would destroy the signal.
"""

from __future__ import annotations

import csv
import itertools
import zipfile
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path

MMOL_TO_MGDL = 18.0182


@dataclass(slots=True)
class Reading:
    dataset_id: str
    source_subject_id: str
    session_id: str
    local_datetime: datetime
    glucose_mg_dl: float
    source_file: str
    source_row: int
    #: False when the source gives only time-of-day and the calendar date was reconstructed.
    date_is_real: bool = True


def _f(x: object) -> float | None:
    try:
        v = float(str(x).strip())
    except (TypeError, ValueError):
        return None
    return v if v == v else None  # reject NaN


def read_big_ideas(root: Path) -> Iterator[Reading]:
    """PhysioNet BIG IDEAs 1.1.2 Dexcom G6 exports.

    The Dexcom export prefixes each file with device/alert metadata rows; only rows whose
    ``Event Type`` is ``EGV`` are sensor glucose readings.
    """
    for path in sorted(root.glob("*/Dexcom_*.csv")):
        subject = path.stem.split("_")[-1]
        with open(path, newline="", encoding="utf-8-sig") as fh:
            for i, row in enumerate(csv.DictReader(fh)):
                if (row.get("Event Type") or "").strip() != "EGV":
                    continue
                ts = (row.get("Timestamp (YYYY-MM-DDThh:mm:ss)") or "").strip()
                val = _f(row.get("Glucose Value (mg/dL)"))
                if not ts or val is None:
                    continue
                yield Reading(
                    "big_ideas", subject, f"big_ideas/{subject}",
                    datetime.fromisoformat(ts), val, str(path), i,
                )


def read_stanford(root: Path) -> Iterator[Reading]:
    """Stanford CGM series from the companion repo (DECISIONS D003), not the smoothed page."""
    path = root / "filtered_cgm_03222026.csv"
    with open(path, newline="") as fh:
        for i, row in enumerate(csv.DictReader(fh)):
            val = _f(row.get("glucose_value"))
            if val is None:
                continue
            subject = row["subject"]
            yield Reading(
                "stanford", subject, f"stanford/{subject}",
                datetime.fromisoformat(row["timestamp"]), val, str(path), i,
            )


def read_shanghai(root: Path, cohort: str = "T2DM") -> Iterator[Reading]:
    """ShanghaiT1DM/T2DM, one workbook per recording visit.

    The filename encodes ``{person}_{visit}_{date}``, which is how person identity and repeat
    visits are separated (blueprint §6.4). Values are mg/dL despite the 15-minute cadence.
    """
    import openpyxl
    import xlrd

    files = [
        p for p in sorted(root.rglob(f"*{cohort}*/*.xls*")) if "Summary" not in p.name
    ]
    for path in files:
        stem = path.stem  # e.g. 2001_1_20201117
        parts = stem.split("_")
        person, visit = (parts[0], parts[1]) if len(parts) >= 2 else (stem, "0")
        rows: list[tuple[object, object]] = []
        if path.suffix == ".xlsx":
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = [(r[0], r[1]) for r in ws.iter_rows(values_only=True)][1:]
            wb.close()
        else:
            book = xlrd.open_workbook(str(path))
            sh = book.sheet_by_index(0)
            for r in range(1, sh.nrows):
                cell = sh.cell(r, 0)
                dt = (
                    xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
                    if cell.ctype == 3
                    else cell.value
                )
                rows.append((dt, sh.cell_value(r, 1)))
        for i, (raw_dt, raw_val) in enumerate(rows):
            val = _f(raw_val)
            if val is None or raw_dt in (None, ""):
                continue
            dt = raw_dt if isinstance(raw_dt, datetime) else datetime.fromisoformat(str(raw_dt))
            yield Reading(
                f"shanghai_{cohort.lower()}", person,
                f"shanghai_{cohort.lower()}/{person}/visit_{visit}", dt, val, str(path), i,
            )


def read_colas(root: Path, anchor: datetime | None = None) -> Iterator[Reading]:
    """Colas iPro recordings.

    The source records **time-of-day only** (``hora``), with no calendar date. The date is
    reconstructed by detecting midnight wraparound and is therefore synthetic: ``date_is_real``
    is False. Time-of-day itself is genuine, so circadian phase is preserved exactly — which is
    what the model actually consumes (§8.5). Recordings are typically ~2 days.
    """
    anchor = anchor or datetime(2000, 1, 1)
    zip_path = next(root.glob("*.zip"), None)
    if zip_path is None:
        return
    with zipfile.ZipFile(zip_path) as zf:
        names = sorted(
            n for n in zf.namelist() if n.lower().endswith(".csv") and "case" in n.lower()
        )
        for name in names:
            case = Path(name).stem.replace("case", "").strip()
            with zf.open(name) as raw:
                text = raw.read().decode("utf-8-sig", errors="replace").splitlines()
            day = 0
            prev: str | None = None
            for i, row in enumerate(csv.DictReader(text)):
                hora = (row.get("hora") or "").strip()
                val = _f(row.get("glucemia"))
                if not hora or val is None:
                    continue
                if prev is not None and hora < prev:
                    day += 1  # wrapped past midnight
                prev = hora
                h, m, s = (int(x) for x in hora.split(":"))
                dt = anchor + timedelta(days=day, hours=h, minutes=m, seconds=s)
                yield Reading(
                    "colas", case, f"colas/{case}", dt, val, f"{zip_path}:{name}", i,
                    date_is_real=False,
                )


def read_hall(root: Path) -> Iterator[Reading]:
    """Hall glucotypes CGM series (supporting file s010, AMENDMENTS A3).

    ``DisplayTime`` is the privacy-shifted local wall clock and is the correct field for
    circadian phase; ``InternalTime`` disagrees by years (§8.5).
    """
    path = root / "S1_cgm_timeseries.tsv"
    with open(path, newline="") as fh:
        for i, row in enumerate(csv.DictReader(fh, delimiter="\t")):
            val = _f(row.get("GlucoseValue"))
            if val is None:
                continue
            subject = row["subjectId"]
            yield Reading(
                "hall", subject, f"hall/{subject}",
                datetime.fromisoformat(row["DisplayTime"]), val, str(path), i,
            )


# CGMacros publishes a 1-minute grid in which both CGM columns have been linearly interpolated
# from their native cadence. Blueprint §19.10 asks for "original raw Dexcom and Libre exports,
# not its interpolated CGM columns", but the 1.0.0 release contains no such export: each subject
# folder holds only meal photographs and this one file. See BLUEPRINT_AMENDMENTS A6 and D016.
CGMACROS_NATIVE_MINUTES = {"Dexcom GL": 5, "Libre GL": 15}
SLOPE_TOLERANCE = 1e-6


def recover_observations(
    minutes: list[float], values: list[float | None]
) -> list[int]:
    """Indices of the original observations behind a linearly interpolated series.

    Linear interpolation is piecewise linear, so the slope is constant between consecutive true
    observations and changes at each one. The observations are therefore recoverable exactly as
    the points where the second difference is non-zero, plus the two endpoints.

    This is not a heuristic reconstruction and is not the same thing as resampling. The claim is
    checkable, and `tests/golden/test_cgmacros.py` checks it on real subjects: re-interpolating
    the recovered points reproduces the published file to floating-point round-off, and the
    recovered timestamps land on the sensor's native grid.

    We do this rather than accept the 1-minute series because the physical observation mask is
    authoritative throughout this project. Reading the interpolated column would report a Dexcom
    window as fully observed when four of every five values were manufactured.
    """
    present = [i for i, v in enumerate(values) if v is not None]
    if len(present) < 3:
        return present
    slopes = []
    for a, b in itertools.pairwise(present):
        dt = minutes[b] - minutes[a]
        slopes.append((values[b] - values[a]) / dt if dt else 0.0)
    keep = [present[0]]
    for i in range(1, len(slopes)):
        if abs(slopes[i] - slopes[i - 1]) > SLOPE_TOLERANCE:
            keep.append(present[i])
    keep.append(present[-1])
    return sorted(set(keep))


def read_cgmacros(root: Path, sensor: str = "Dexcom GL") -> Iterator[Reading]:
    """CGMacros 1.0.0. Evaluation only — Lane E pending Q4.

    Yields only the recovered native observations, never the interpolated filler. The two sensors
    are read separately at their own cadence, as §19.10 requires; they are paired by subject at
    split time rather than merged into one series.
    """
    if sensor not in CGMACROS_NATIVE_MINUTES:
        raise ValueError(f"unknown CGMacros sensor {sensor!r}")
    archive = root / "1.0.0/CGMacros_dateshifted365.zip"
    with zipfile.ZipFile(archive) as z:
        names = sorted(
            n for n in z.namelist()
            if n.endswith(".csv") and "/CGMacros-" in n and "DataDictionary" not in n
        )
        for name in names:
            subject = name.rsplit("/", 1)[-1].removesuffix(".csv").removeprefix("CGMacros-")
            with z.open(name) as handle:
                rows = list(csv.DictReader(line.decode("utf-8-sig") for line in handle))
            stamps: list[datetime] = []
            values: list[float | None] = []
            for row in rows:
                try:
                    stamps.append(datetime.fromisoformat(row["Timestamp"].strip()))
                except (KeyError, ValueError):
                    stamps.append(None)  # type: ignore[arg-type]
                values.append(_f(row.get(sensor)))
            usable = [i for i, t in enumerate(stamps) if t is not None]
            if not usable:
                continue
            origin = stamps[usable[0]]
            minutes = [
                (stamps[i] - origin).total_seconds() / 60.0 if stamps[i] else float("nan")
                for i in range(len(stamps))
            ]
            for i in recover_observations(minutes, values):
                if stamps[i] is None or values[i] is None:
                    continue
                yield Reading(
                    dataset_id="cgmacros",
                    source_subject_id=subject,
                    session_id=f"cgmacros/{subject}/{sensor.split()[0].lower()}",
                    local_datetime=stamps[i],
                    glucose_mg_dl=values[i],
                    source_file=name,
                    source_row=i,
                )


READERS = {
    "big_ideas": read_big_ideas,
    "stanford": read_stanford,
    "shanghai_t2dm": read_shanghai,
    "colas": read_colas,
    "hall": read_hall,
    "cgmacros": read_cgmacros,
}
